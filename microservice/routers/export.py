import io
import json
import uuid

from domain.errors import ProcesoNoEncontradoError
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from infrastructure.persistence.repositorio_proceso import PostgresProcesoRepo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

router = APIRouter()
repo = PostgresProcesoRepo()

HEADER_FILL = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALTA_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
MEDIA_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

COLUMNAS = [
    ("NIT", 15),
    ("Razon Social", 35),
    ("CIIU", 10),
    ("Clasificacion", 22),
    ("Score Unificado", 16),
    ("SRF", 10),
    ("Nivel Riesgo", 14),
    ("Hallazgos", 60),
    ("Explicacion IA", 80),
]


@router.get("/proceso/{proceso_id}/export")
async def exportar_resultados(
    proceso_id: uuid.UUID,
    formato: str = Query("xlsx", regex="^xlsx$"),
):
    proceso = await repo.obtener_proceso(proceso_id)
    if not proceso:
        raise ProcesoNoEncontradoError(str(proceso_id))

    total, rows = await repo.listar_proceso_detalle(
        proceso_id=proceso_id,
        page=1,
        page_size=10000,
        ordenar_por="mcp_score",
        direccion="desc",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Campana"

    for col_idx, (nombre, ancho) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = ancho

    for row_idx, r in enumerate(rows, 2):
        hallazgos_raw = r.get("hallazgos") or []
        if isinstance(hallazgos_raw, str):
            hallazgos_raw = json.loads(hallazgos_raw)
        hallazgos_text = "; ".join(
            f"{h.get('tipo', '')} ({h.get('severidad', '')})" for h in hallazgos_raw
        )

        valores = [
            r.get("nit", ""),
            r.get("razon_social", ""),
            r.get("ciiu", ""),
            r.get("clasificacion", ""),
            r.get("mcp_score") or 0,
            r.get("srf_total") or 0,
            r.get("nivel_riesgo", ""),
            hallazgos_text,
            (r.get("explicacion_ia") or "")[:500],
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        nivel = r.get("nivel_riesgo", "")
        if nivel == "ALTO":
            for col_idx in range(1, len(COLUMNAS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALTA_FILL
        elif nivel == "MEDIO":
            for col_idx in range(1, len(COLUMNAS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = MEDIA_FILL

    ws_resumen = wb.create_sheet("Resumen", 0)
    resumen_data = [
        ("Proceso ID", str(proceso_id)),
        ("Estado", proceso.get("estado", "")),
        ("Total NITs", proceso.get("total_nits", 0)),
        ("Candidatos", proceso.get("candidatos", 0)),
        ("Omisos", proceso.get("omisos", 0)),
        ("Inexactos", proceso.get("inexactos", 0)),
    ]
    for row_idx, (label, value) in enumerate(resumen_data, 1):
        ws_resumen.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_resumen.cell(row=row_idx, column=2, value=value)
    ws_resumen.column_dimensions["A"].width = 20
    ws_resumen.column_dimensions["B"].width = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"fiscalia_campana_{proceso_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
